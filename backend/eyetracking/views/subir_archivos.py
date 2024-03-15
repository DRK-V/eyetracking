# Importing necessary libraries
import os
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db import connection
from datetime import datetime 
import numpy as np 
import pandas as pd
import time
import openpyxl

@csrf_exempt
def subir_archivos(request):
    print('SUBIENDO ARCHIVO ||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||||')
    if request.method == 'POST':
        try:
            archivos = request.FILES.getlist('archivos')
            proyecto_id = request.session.get('proyecto_id', None)
            start_time = time.time()
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT fecha_creacion FROM proyecto WHERE id_proyecto = (%s)",
                    [proyecto_id]
                )
                fecha_proyecto_str = cursor.fetchone()[0]

            if fecha_proyecto_str:
                fecha_proyecto_formateada = fecha_proyecto_str.strftime('%Y%m%d')
                print(f'Fecha obtenida al subir archivo para insertar el excel dentro de ARCHIVOS_SUBIDOS: {fecha_proyecto_str}')
                
                carpeta_proyecto = f'proyecto{proyecto_id}_{fecha_proyecto_formateada}'
                carpeta_media = os.path.join(settings.MEDIA_ROOT, carpeta_proyecto)
                carpeta_videos = os.path.join(carpeta_media, 'videos')
                carpeta_imagenes = os.path.join(carpeta_media, 'imagenes')
                carpeta_archivos_subidos = os.path.join(carpeta_media, 'archivos_subidos')

                print("Archivos recibidos:")
                for archivo in archivos:
                    print(archivo.name)
                    if archivo.content_type.startswith('video'):
                        ruta_archivo = os.path.join(carpeta_videos, archivo.name)
                    elif archivo.content_type.startswith('image'):
                        ruta_archivo = os.path.join(carpeta_imagenes, archivo.name)
                    else:
                        ruta_archivo = os.path.join(carpeta_archivos_subidos, archivo.name)

                    print(f' Ruta archivo la cual se guardara el archivo: {ruta_archivo}')
                    if os.path.exists(ruta_archivo):
                        with open(ruta_archivo, 'wb') as f:
                            for chunk in archivo.chunks():
                                f.write(chunk)
                    else:
                        os.makedirs(carpeta_media, exist_ok=True)
                        with open(ruta_archivo, 'wb') as f:
                            for chunk in archivo.chunks():
                                f.write(chunk)

                # Iteración para el archivo 'metricas.xlsx'
                for archivo in archivos:
                    if archivo.name == 'metricas.xlsx':
                        df = pd.read_excel(archivo)
                        df.replace({np.nan: None}, inplace=True)  # Reemplaza nan con None

                        columnas_deseadas = ['Recording timestamp', 'Sensor', 'Genero', 'Edad', 'Pupil diameter left', 'Pupil diameter right','Recording name']

                        data = [{columna: valor_campo[index] for index, columna in enumerate(columnas_deseadas)} for valor_campo in zip(*[df[col] for col in columnas_deseadas])]

                        insertar_en_metricas(request, data)
                        end_time = time.time()  # Detiene el temporizador
                        elapsed_time = end_time - start_time
                        #print(f"Inserción completada. Se insertaron {len(data)} registros. Tiempo total: {elapsed_time} segundos")

                # Iteración para el archivo 'interes_e_impacto.xlsx'
                for archivo in archivos:
                    if archivo.name == 'interes_e_impacto.xlsx':
                        process_interes_impacto_file(request, archivo)

                return JsonResponse({'message': 'Archivos cargados exitosamente'})        
        except Exception as e:
            print(f"Error: {e}")
            return JsonResponse({'error': str(e)}, status=400)

    return render(request, 'componentes/home/upload_files.html')

def insertar_en_metricas(request, data):
    proyecto_id = request.session.get('proyecto_id', None)

    if not proyecto_id:
        print(f' el id del proyecto no es valido {proyecto_id}')
        return 

    with connection.cursor() as cursor:
        cursor.executemany(
            "INSERT INTO metricas (codigo_proyecto, recording_timestamp, sensor, genero, edad, pupil_diameter_left, pupil_diameter_right, recording_name) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
            [(proyecto_id, item['Recording timestamp'], item['Sensor'], item['Genero'], item['Edad'], item['Pupil diameter left'], item['Pupil diameter right'], item['Recording name']) for item in data]
        )

        print(f"Inserción en metricas: {len(data)} filas insertadas")

def organizar_datos(datos_base):
    datos_organizados = {}

    for key, value in datos_base.items():
        # Manejar casos donde la división no genera exactamente dos partes
        try:
            tipo, numero = key.split(' ', 1)
        except ValueError:
            print(f"Error al dividir la clave: {key}")
            continue

        if tipo not in datos_organizados:
            datos_organizados[tipo] = {}

        datos_organizados[tipo][numero] = value

    return datos_organizados




def process_interes_impacto_file(request, archivo):
    proyecto_id = request.session.get('proyecto_id', None)

    if not proyecto_id:
        print(f'El id del proyecto no es válido: {proyecto_id}')
        return 

    datos_base_interes = {}  # Diccionario para almacenar los datos de la hoja "Fixation count incl 0"
    datos_base_impacto = {}  # Diccionario para almacenar los datos de la hoja "Time to first Fixation"

    try:
        # Cargar el libro de Excel
        wb = openpyxl.load_workbook(archivo)

        # Especificar los nombres correctos de las hojas
        sheet_names = ['Fixation count incl 0', 'Time to first Fixation']

        for sheet_name in sheet_names:
            # Verificar si la hoja existe en el libro
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Nombres que estamos buscando en cada celda de la fila
                target_names_interes = [
                    'Number of fixations in AOI (including zeroes)',
                    'Participant',
                    'Edad',
                    'Genero',
                    'Marco',
                    'Patas',
                    'Puente',
                    'Average',
                    'Median',
                    'Sum',
                    'Total Time of Interest Fixation Count',
                    'Total Time of Interest Duration',
                    'Total Recording Duration',
                ]

                target_names_impacto = [
                    'Time to first fixation in AOI',
                    'Participant',
                    'Edad',
                    'Genero',
                    'Marco',
                    'Patas',
                    'Puente',
                    'Average',
                    'Median',
                    'Count',
                    'Total Time of Interest Duration',
                    'Total Recording Duration',
                ]

                previous_row = None  # Variable para almacenar la fila anterior
                current_table_number = 0  # Inicializar el número de tabla

                datos_base = datos_base_interes if sheet_name == 'Fixation count incl 0' else datos_base_impacto
                target_names = target_names_interes if sheet_name == 'Fixation count incl 0' else target_names_impacto

                for current_row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if all(name in str(cell) for name, cell in zip(target_names, row)):
                        if previous_row is not None:
                            # Filtrar los valores nulos o vacíos en la referencia
                            filtered_previous_row = [str(cell) for cell in previous_row if cell is not None]
                            datos_base[f"Referencia {current_table_number}"] = filtered_previous_row

                        current_table_number += 1
                        current_row_in_table = 1
                        rows_count = 0

                        filtered_row = [str(cell) if cell is not None else "None" for cell in row]  # Convertir valores None a "None"
                        datos_base[f"Columnas {current_table_number}"] = filtered_row[:12]  # Tomar solo los primeros 12 valores

                        for next_row in sheet.iter_rows(min_row=current_row_number + 1, values_only=True):
                            if not any(next_row) or any(next_row[0].lower() == keyword.lower() for keyword in ["average", "count", "variance", "standard deviation (n-1)"]):
                                break

                            # Filtrar los valores nulos o vacíos en las filas de datos
                            filtered_next_row = [str(cell) if cell is not None else "None" for cell in next_row[:12]]
                            datos_base[f"Fila {current_row_in_table} de la Tabla {current_table_number}"] = filtered_next_row
                            current_row_in_table += 1
                            rows_count += 1

                        # print(f"La Tabla número {current_table_number} tiene {rows_count} filas.")

                    previous_row = row

        # Llamar a las funciones correspondientes para insertar en las tablas
        insertar_en_interes(request, proyecto_id, datos_base_interes)
        insertar_en_impacto(request, proyecto_id, datos_base_impacto)

    except Exception as e:
        print(f"Error al procesar el archivo: {e}")


def insertar_en_interes(request, proyecto_id, datos_base):
    try:
        with connection.cursor() as cursor:
            for key, values in datos_base.items():
                if key.startswith('Referencia'):
                    ref = values[0]
                    codigo_proyecto = proyecto_id
                elif key.startswith('Columnas'):
                    columnas = values
                elif key.startswith('Fila'):
                    # Asegurémonos de que la fila tenga suficientes valores
                    if len(values) >= len(columnas):
                        # Crear un diccionario con valores y columnas
                        values_dict = {col: value if value is not None and value != 'None' else None for col, value in zip(columnas, values)}
                        values_dict.update({'ref': ref, 'codigo_proyecto': codigo_proyecto})
                        
                        # Convertir valores vacíos a None
                        for col, value in values_dict.items():
                            if value == '':
                                values_dict[col] = None
                      
                        # Asegurémonos de que todas las columnas necesarias están presentes
                        if all(col in values_dict for col in [
                            'Participant',
                            'Edad',
                            'Genero',
                            'Marco',
                            'Patas',
                            'Puente',
                            'Average',
                            'Median',
                            'Sum'
                        ]):
                            cursor.execute(
                                "INSERT INTO interes (codigo_proyecto, ref, "
                                "participant, edad, genero, marco, patas, puente, "
                                "average, median, sum) "
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                                (
                                    codigo_proyecto,
                                    values_dict['ref'],
                                    values_dict['Participant'],
                                    values_dict['Edad'],
                                    values_dict['Genero'],
                                    values_dict['Marco'],
                                    values_dict['Patas'],
                                    values_dict['Puente'],
                                    values_dict['Average'],
                                    values_dict['Median'],
                                    values_dict['Sum']
                                )
                            )

    except Exception as e:
        print(f"Error al insertar en la tabla interes: {e}")





def insertar_en_impacto(request, proyecto_id, datos_base):
    try:
        with connection.cursor() as cursor:
            for key, values in datos_base.items():
                if key.startswith('Referencia'):
                    ref = values[0]
                    codigo_proyecto = proyecto_id
                elif key.startswith('Columnas'):
                    columnas = values
                elif key.startswith('Fila'):
                    # Asegurémonos de que la fila tenga suficientes valores
                    if len(values) >= len(columnas):
                        # Crear un diccionario con valores y columnas
                        values_dict = {col: value if value is not None and value != 'None' else None for col, value in zip(columnas, values)}
                        values_dict.update({'ref': ref, 'codigo_proyecto': codigo_proyecto})

                        # Convertir valores vacíos a None
                        for col, value in values_dict.items():
                            if value == '':
                                values_dict[col] = None
                      
                        # Asegurémonos de que todas las columnas necesarias están presentes
                        if all(col in values_dict for col in [
                                'Participant',
                                'Edad',
                                'Genero',
                                'Marco',
                                'Patas',
                                'Puente',
                                'Average',
                                'Median',
                                'Count'
                            ]):
                                cursor.execute(
                                    "INSERT INTO impacto (codigo_proyecto, ref, "
                                   "participant, edad, genero, marco, patas, puente, "
                                    "average, median, count) "
                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                                    (
                                        codigo_proyecto,
                                        values_dict['ref'],
                                        values_dict['Participant'],
                                        values_dict['Edad'],
                                        values_dict['Genero'],
                                        values_dict['Marco'],
                                        values_dict['Patas'],
                                        values_dict['Puente'],
                                        values_dict['Average'],
                                        values_dict['Median'],
                                        values_dict['Count'],
                                    )
                                )
                               
    except Exception as e:
        print(f"Error al insertar en la tabla impacto: {e}")
